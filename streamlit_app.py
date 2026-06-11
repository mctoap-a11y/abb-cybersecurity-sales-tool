import streamlit as st
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import io

# PAGE CONFIG
st.set_page_config(page_title="ABB Cybersecurity Sales Tool", page_icon="shield", layout="wide")

# PRODUCTS
PRODUCTS = {
    "1": {"name": "Risk Assessment", "category": "Professional Services", "short": "IEC 62443 compliant risk assessments"},
    "2": {"name": "CSWP", "category": "Endpoint Protection", "short": "Application whitelisting for OT"},
    "3": {"name": "SUS", "category": "Patch Management", "short": "Managed OT-safe patch delivery"},
    "4": {"name": "Backups (Acronis)", "category": "Business Continuity", "short": "OT backup and disaster recovery"},
    "5": {"name": "Allowlisting (Trellix)", "category": "Endpoint Protection", "short": "Enterprise application control"},
    "6": {"name": "SIEM Connector", "category": "Security Monitoring", "short": "OT event forwarding to SIEM"},
    "7": {"name": "SIEM", "category": "Security Operations", "short": "Full OT SIEM deployment"},
    "8": {"name": "Network Monitoring", "category": "OT Network Security", "short": "Passive OT network visibility"},
    "9": {"name": "MFA", "category": "Identity & Access", "short": "Strong authentication for OT"},
}

PRICES = {
    "cswp": {"name": "CSWP Only", "price": 197.10, "currency": "EUR", "applies": ["server", "workstation"]},
    "cswp_sus": {"name": "CSWP + SUS", "price": 210.00, "currency": "EUR", "applies": ["server", "workstation"]},
    "acronis_srv": {"name": "Acronis Backup - Server", "price": 788.00, "currency": "GBP", "applies": ["server"]},
    "acronis_ws": {"name": "Acronis Backup - Workstation", "price": 68.00, "currency": "GBP", "applies": ["workstation"]},
    "trellix_srv": {"name": "Trellix - Server", "price": 170.00, "currency": "USD", "applies": ["server"]},
    "trellix_ws": {"name": "Trellix - Workstation", "price": 21.50, "currency": "USD", "applies": ["workstation"]},
    "siem": {"name": "SIEM Connector", "price": 157.50, "currency": "EUR", "applies": ["server", "workstation"]},
}

# HEADER
st.markdown("<h1 style='color: #FF000F;'>ABB Cybersecurity Sales Intelligence Platform</h1>", unsafe_allow_html=True)
st.write("Your complete OT/ICS cybersecurity sales companion - Modern Web Version")

# TABS
tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
    "Products", "IEC 62443", "NIS2", "UK CSRB", "CNI vs Others", "Price Estimator", "Sales Resources"
])

# PRODUCTS TAB
with tab1:
    st.subheader("ABB Cybersecurity Products")
    product_names = [PRODUCTS[k]["name"] for k in PRODUCTS.keys()]
    selected = st.selectbox("Select Product:", product_names)
    selected_key = [k for k in PRODUCTS.keys() if PRODUCTS[k]["name"] == selected][0]
    product = PRODUCTS[selected_key]
    
    st.write(f"**Category:** {product['category']}")
    st.write(f"**Description:** {product['short']}")

# IEC 62443 TAB
with tab2:
    st.subheader("IEC 62443 - Industrial Automation Security")
    st.write("International standard for OT/ICS cybersecurity with 4 security levels (SL 1-4)")
    
    df = pd.DataFrame({
        "Level": ["SL 1", "SL 2", "SL 3", "SL 4"],
        "Name": ["Basic", "Moderate", "High", "Critical"],
        "Use Case": ["Low risk OT", "General industry", "CNI/NIS2 Essential", "State-sponsored attacks"]
    })
    st.dataframe(df, use_container_width=True)

# NIS2 TAB
with tab3:
    st.subheader("NIS2 - EU Cybersecurity Directive")
    st.write("Fines up to EUR 10M for Essential entities, EUR 7M for Important entities")
    
    df = pd.DataFrame({
        "Article": ["21.2a", "21.2b", "21.2c", "21.2d", "21.2e", "21.2j"],
        "Requirement": ["Risk analysis", "Incident handling", "Business continuity", "Supply chain", "Network security", "MFA"]
    })
    st.dataframe(df, use_container_width=True)

# UK CSRB TAB
with tab4:
    st.subheader("UK CSRB - Cyber Security & Resilience Bill")
    st.write("UK critical infrastructure protection requirements")
    
    df = pd.DataFrame({
        "Sector": ["Energy", "Water", "Transport", "Health", "Finance"],
        "SL Target": ["SL 3-4", "SL 3", "SL 2-3", "SL 2-3", "SL 3"]
    })
    st.dataframe(df, use_container_width=True)

# CNI VS OTHERS
with tab5:
    st.subheader("Sliding Scale: CNI vs Non-CNI")
    
    c1, c2, c3 = st.columns(3)
    
    with c1:
        st.markdown("**CNI Designated**")
        st.write("- SL 3+ compliance")
        st.write("- 24-hour reporting")
        st.write("- 24/7 monitoring")
        st.write("- Mandatory MFA")
    
    with c2:
        st.markdown("**NIS2-Regulated**")
        st.write("- SL 2 recommended")
        st.write("- 72-hour reporting")
        st.write("- Endpoint protection")
        st.write("- Backups required")
    
    with c3:
        st.markdown("**Non-Regulated**")
        st.write("- No mandate")
        st.write("- Insurance-driven")
        st.write("- Cyber Essentials")
        st.write("- Recommended")

# PRICE ESTIMATOR TAB
with tab6:
    st.subheader("Price Estimator")
    st.warning("Internal transfer prices only - add margin for sell price")
    
    c1, c2 = st.columns(2)
    
    with c1:
        servers = st.number_input("Servers:", 0)
        workstations = st.number_input("Workstations:", 0)
    
    with c2:
        st.write("Select Products:")
        selected_products = {}
        for key, price in PRICES.items():
            selected_products[key] = st.checkbox(f"{price['name']} - {price['currency']} {price['price']:.2f}")
    
    if st.button("Calculate Estimate"):
        selected = [k for k, v in selected_products.items() if v]
        
        if not selected:
            st.error("Select at least one product!")
        else:
            results = []
            costs = {}
            
            for key in selected:
                price_data = PRICES[key]
                
                if "server" in price_data["applies"] and "workstation" in price_data["applies"]:
                    nodes = servers + workstations
                elif "server" in price_data["applies"]:
                    nodes = servers
                else:
                    nodes = workstations
                
                total = price_data["price"] * nodes
                currency = price_data["currency"]
                
                if currency not in costs:
                    costs[currency] = 0
                costs[currency] += total
                
                results.append({
                    "Product": price_data["name"],
                    "Nodes": nodes,
                    "Price": f"{price_data['price']:.2f}",
                    "Total": f"{total:,.2f}",
                    "Ccy": currency
                })
            
            st.dataframe(pd.DataFrame(results), use_container_width=True)
            
            st.markdown("### Summary")
            for currency, total in sorted(costs.items()):
                st.success(f"**{currency}: {currency} {total:,.2f}/year**")
            
            # EXCEL EXPORT
            wb = Workbook()
            ws = wb.active
            
            ws['A1'] = "ABB Cybersecurity - Price Estimate"
            ws['A1'].font = Font(bold=True, size=14, color="FF000F")
            
            headers = ["Product", "Nodes", "Price/Node", "Total", "Currency"]
            for col, h in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col)
                cell.value = h
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="FF000F", end_color="FF000F", fill_type="solid")
            
            for row, r in enumerate(results, 4):
                ws.cell(row=row, column=1).value = r["Product"]
                ws.cell(row=row, column=2).value = r["Nodes"]
                ws.cell(row=row, column=3).value = r["Price"]
                ws.cell(row=row, column=4).value = r["Total"]
                ws.cell(row=row, column=5).value = r["Ccy"]
            
            ws.column_dimensions['A'].width = 35
            
            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)
            
            st.download_button(
                "Download Estimate (Excel)",
                buf,
                f"ABB_Estimate_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

# SALES RESOURCES TAB
with tab7:
    st.subheader("Sales Resources")
    
    t1, t2, t3, t4, t5 = st.tabs(["Threats", "Regulatory", "ROI", "ABB Differentiators", "Discovery"])
    
    with t1:
        st.markdown("**Threat Reality:** OT attacks up 140% YoY. Avg incident: GBP 2.3M. Ransomware downtime: 23 days.")
    
    with t2:
        st.markdown("**Regulatory:** NIS2 fines EUR 10M. UK CSRB coming. IEC 62443 mandatory in contracts.")
    
    with t3:
        st.markdown("**ROI:** Insurance down 15-30%. Downtime cost: GBP 8K-20K/hour. CSWP payback: 3-6 months.")
    
    with t4:
        st.markdown("**ABB:** OT-native. Works on legacy OS. IEC 62443 aligned. Single vendor for security+automation.")
    
    with t5:
        st.markdown("**Questions:** Are you NIS2 Essential? Last risk assessment? OT patch process? Recovery plan?")

st.markdown("---")
st.write("ABB Cybersecurity Sales Intelligence Platform v3.1 | peter.mctoal@gb.abb.com")